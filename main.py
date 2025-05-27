# /// script
# dependencies = [
#   "cryptography==44.0.2",
#   "dnspython==2.7.0",
#   "edgegrid-python==2.0.0",
#   "ipython==9.0.2",
#   "numpy==2.2.4",
#   "openpyxl==3.1.5",
#   "pandas==2.2.3",
#   "python-dateutil==2.9.0.post0",
#   "python-pptx==1.0.2",
#   "requests==2.32.3",
#   "tqdm==4.67.1",
# ]
# requires-python = "==3.13.1"
# ///

import copy
import csv
import ipaddress
import json
import os
import re
import time
import subprocess
from datetime import date, datetime
import concurrent.futures

import dateutil
import openpyxl
import pandas as pd
from dns import resolver
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from tqdm import tqdm
import shutil
import os

from mylib import (
    extract_behaviors,
    extract_cpcodes,
    extract_im_cpcodes,
    extract_origin_hostnames,
    extract_sr_ssmaps,
    get_cps_enrollment_id,
    get_property_activations,
    get_property_hostnames,
    get_property_info,
    get_property_json,
    get_switch_keys,
    get_traffic_stats,
    list_available_behaviors,
    list_cert_deployments,
    list_cpcodes,
    list_custom_behaviors,
    list_custom_overrides,
    list_edgehostnames,
    list_groups,
    list_properties,
    parse_certificate,
    get_cert,
    extract_row_of_dataframe,
)

# If you have a pre-defined account list, write the list, e.g., ["account1", "account2"].
#ACCOUNTS = []


# If behaviors listed in CRITICAL_BEHAVIORS / RECOMMENDED_BEHAVIORS are not included in the properties, they will be listed in criticalBehaviors / recomendedBehaviors columns in the output.
CRITICAL_BEHAVIORS = {
    "Site_Accel": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
        "persistentConnection",
    ],
    "SPM": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
        "persistentConnection",
    ],
    "Fresca": [
        "enhancedAkamaiProtocol",
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
        "persistentConnection",
    ],
    "Adaptive_Media_Delivery": [
        "tieredDistribution",
        "caching",
    ],
    "Download_Delivery": [
        "largeFileOptimization",
        "tieredDistribution",
        "caching",
    ],
    "Obj_Delivery": [
        "tieredDistribution",
        "caching",
    ],
    "HTTP_Content_Del": [
        "tieredDistribution",
        "caching",
    ],
    "HTTP_Downloads": [
        "tieredDistribution",
        "caching",
    ],
    "Progressive_Media": [
        "tieredDistribution",
        "caching",
    ],
    "Aqua_Mobile": [
        "gzipResponse",
        "tieredDistribution",
    ],
    "Dynamic_Site_Del": [
        "gzipResponse",
        "tieredDistribution",
    ],
    "Site_Del": [
        "gzipResponse",
        "tieredDistribution",
    ],
    "RM": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
    ],
    "IoT": [
        "sureRoute",
    ],
    "Mobile_Accel": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
    ],
    "Obj_Caching": [
        "tieredDistribution",
        "caching",
    ],
    "Rich_Media_Accel": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
    ],
    "Alta": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
    ],
    "Web_App_Accel": [
        "gzipResponse",
        "sureRoute",
        "tieredDistribution",
    ],
    "Site_Defender": [
        "gzipResponse",
        "persistentConnection",
    ],
}

RECOMMENDED_BEHAVIORS = {
    "Site_Accel": [
        "failAction",
        "caching",
        "report",
        "datastream",
        "brotli",
        "dnsPrefresh",
        "siteShield",
    ],
    "SPM": [
        "adaptiveAcceleration",
        "failAction",
        "caching",
        "report",
        "datastream",
        "brotli",
        "dnsPrefresh",
        "siteShield",
    ],
    "Fresca": [
        "adaptiveAcceleration",
        "failAction",
        "caching",
        "report",
        "datastream",
        "brotli",
        "dnsPrefresh",
        "siteShield",
    ],
    "Adaptive_Media_Delivery": [
        "edgeScape",
        "modifyOutgoingResponseHeader",
        "segmentedContentProtection",
        "cacheKeyQueryParams",
        "originFailureRecoveryMethod",
        "dynamicThroughtputOptimization",
        "originFailureRecoveryMethod",
        "manifestPersonalization",
    ],
    "Download_Delivery": [
        "modifyOutgoingResponseHeader",
        "cacheKeyQueryParams",
        "removeVary",
        "report",
        "datastream",
        "dynamicThroughtputOptimization",
    ],
    "Obj_Delivery": [
        "modifyOutgoingResponseHeader",
        "cacheKeyQueryParams",
        "removeVary",
        "report",
        "datastream",
        "dynamicThroughtputOptimization",
    ],
    "HTTP_Content_Del": [
        "deviceCharacteristicCacheId",
        "deviceCharacteristicHeader",
        "largeFileOptimization",
        "gzipResponse",
        "modifyOutgoingResponseHeader",
        "cacheKeyQueryParams",
        "removeVary",
    ],
    "HTTP_Downloads": [
        "largeFileOptimization",
        "modifyOutgoingResponseHeader",
        "cacheKeyQueryParams",
        "removeVary",
        "report",
        "datastream",
    ],
    "Progressive_Media": [
        "modifyOutgoingResponseHeader",
        "cacheKeyQueryParams",
        "removeVary",
        "report",
        "datastream",
    ],
    "Aqua_Mobile": [
        "deviceCharacteristicCacheId",
        "deviceCharacteristicHeader",
        "failAction",
        "caching",
    ],
    "Dynamic_Site_Del": [
        "failAction",
        "caching",
    ],
    "Site_Del": [
        "deviceCharacteristicCacheId",
        "deviceCharacteristicHeader",
        "caching",
    ],
    "RM": [
        "adaptiveAcceleration",
        "failAction",
        "caching",
    ],
    "IoT": [],
    "Mobile_Accel": [
        "caching",
    ],
    "Obj_Caching": [
        "modifyOutgoingResponseHeader",
        "cacheKeyQueryParams",
        "removeVary",
        "report",
        "datastream",
    ],
    "Rich_Media_Accel": [
        "deviceCharacteristicCacheId",
        "deviceCharacteristicHeader",
        "failAction",
        "caching",
    ],
    "Alta": [
        "deviceCharacteristicCacheId",
        "deviceCharacteristicHeader",
        "failAction",
        "caching",
    ],
    "Web_App_Accel": [
        "deviceCharacteristicCacheId",
        "deviceCharacteristicHeader",
        "failAction",
        "caching",
    ],
    "Site_Defender": [
        "failAction",
        "caching",
        "report",
        "datastream",
        "dnsPrefresh",
        "siteShield",
    ],
}

# Behavior columns will be ordered based on COLUMN_ORDER. Behaviors not in COLUMN_ORDER will follow in alphabetical order.
COLUMN_ORDER = [
    "cpCode",
    "sureRoute",
    "tieredDistribution",
    "siteShield",
    "allHttpInCacheHierarchy",
    "allowPost",
    "allowOptions",
    "allowPut",
    "allowPatch",
    "allowDelete",
    "origin",
    "persistentConnection",
    "healthDetection",
    "timeout",
    "readTimeout",
    "failAction",
    "caching",
    "cacheKeyQueryParams",
    "cacheKeyIgnoreCase",
    "cacheError",
    "downstreamCache",
    "removeVary",
    "largeFileOptimization",
    "imageManager",
    "imageManagerVideo",
    "httpStrictTransportSecurity",
    "enhancedAkamaiProtocol",
    "http2",
    "http3",
    "allowTransferEncoding",
    "gzipResponse",
    "brotli",
    "dnsPrefresh",
    "adaptiveAcceleration",
    "dynamicThroughtputOptimization",
    "mPulse",
    "report",
    "logCustom",
    "datastream",
    "modifyIncomingRequestHeader",
    "modifyOutgoingRequestHeader",
    "modifyIncomingResponseHeader",
    "modifyOutgoingResponseHeader",
    "redirect",
    "redirectplus",
    "constructResponse",
    "denyAccess",
    "advanced",
    "customBehavior",
    "advancedOverride",
    "customOverride",
]

PROPERTY_VER = "production"
N_CHUNK = 5



def load_accounts_from_csv(file_path):
    import csv
    accounts = []
    try:
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if row and row[0].strip().lower() != "account":
                    accounts.append(row[0].strip())
    except Exception as e:
        print(f"❌ Error reading accounts CSV: {e}")
    return accounts

# Prompt user for CSV file path
csv_path = input("Enter path to account CSV file (e.g., accounts.csv): ").strip()
ACCOUNTS = load_accounts_from_csv(csv_path)


def chunked_list(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i : i + size]


def property_analysis(
    switch_key,
    property_name,
    df_enroll,
    df_custom_behaviors=None,
    df_custom_overrides=None,
    target_ver=PROPERTY_VER,
    traffic_report=False,
    server_cert=False,
):
    def get_cname_chain(domain):
        cname_chain = []

        try:
            answer = resolver.resolve(domain, "CNAME")
            for rdata in answer:
                cname = rdata.target.to_text()
                ttl = answer.rrset.ttl
                cname_chain.append((cname.strip("."), ttl))
                cname_chain.extend(get_cname_chain(cname.strip(".")))
        except:
            pass

        return cname_chain

    def check_xml(xml, features):
        results = []
        if type(xml) is str:
            if all([word in xml for word in ["list-names", "NL_"]]):
                results.append("Network List")
            for feature in features:
                if feature in xml:
                    results.append(feature)

        return results

    retry_counter = 2
    while retry_counter >= 0:
        try:
            df_property = get_property_info(
                switch_key,
                propertyname=property_name,
                target_ver=PROPERTY_VER,
            )
            property_data = {
                "accountName": account,
                "accountSwitchKey": switch_key,
            }
            property_data.update(df_property.iloc[0].to_dict())
            comments = {
                key: []
                for key in [
                    "hostnames",
                    "cnameChain",
                    "edgeHostnames",
                    "slots",
                    "realCertificates",
                    "originHostnames",
                    "cpCodes",
                ]
            }

            contract_id = property_data["contractId"]
            group_id = property_data["groupId"]
            property_id = property_data["propertyId"]
            version = property_data["propertyVersion"]

            # Version Note
            property_data["versionNote"] = (
                property_data.pop("note") if "note" in property_data.keys() else None
            )

            # Activation info
            try:
                df_activations = get_property_activations(
                    switch_key,
                    property_id,
                    network=PROPERTY_VER,
                    version=version,
                )
                property_data["activatedDate"] = df_activations.at[0, "updateDate"]
                property_data["activationNote"] = (
                    df_activations.at[0, "note"]
                    if "note" in df_activations.columns
                    else None
                )
            except:
                property_data["activatedDate"] = None
                property_data["activationNote"] = None

            for col in [
                "updatedByUser",
                "updatedDate",
                "versionNote",
                "stagingStatus",
                "productionStatus",
                "activatedDate",
                "activationNote",
            ]:
                property_data[col] = (
                    property_data.pop(col) if col in property_data.keys() else None
                )

            # Property hotnames
            df_hostnames = get_property_hostnames(switch_key, property_id, version)
            property_data["hostnames"] = "\n".join(list(df_hostnames["cnameFrom"]))

            if any(
                [
                    zone in property_data["hostnames"]
                    for zone in [
                        "edgekey.net",
                        "edgesuite.net",
                        "edgekey-staging.net",
                        "edgesuite-staging.net",
                    ]
                ]
            ):
                comments["hostnames"].append(
                    "Edge hostname is configured as property hostname"
                )

            # CNAME records on Property hostnames
            cnames = []
            for hostname in list(df_hostnames["cnameFrom"]):
                records = get_cname_chain(hostname)
                if records:
                    if any([int(ttl) >= 86400 for rdata, ttl in records if ttl]):
                        comments["cnameChain"].append(f"TTL for CNAME is too long")
                    records = [f"{rdata}(TTL:{ttl})" for rdata, ttl in records]
                    cnames.append(" > ".join([hostname] + records))

                    cnamed = False
                    for record in [hostname] + records:
                        if any(
                            [
                                (edgehost in record)
                                for edgehost in [
                                    "edgekey.net",
                                    "edgesuite.net",
                                    "akamaized.net",
                                    "edgekey-staging.net",
                                    "edgesuite-staging.net",
                                    "akamaized-staging.net",
                                ]
                            ]
                        ):
                            cnamed = True
                            break
                    if not cnamed:
                        comments["cnameChain"].append(
                            f"{hostname} is not CNAMED to Akamai"
                        )

            property_data["cnameChain"] = "\n".join(cnames) if cnames else None

            # Edge hostnames
            zipped_hostnames = [
                (host, edge_host)
                for host, edge_host in zip(
                    df_hostnames["cnameFrom"], df_hostnames["cnameTo"]
                )
                if edge_host
            ]
            slots = []

            if zipped_hostnames:
                details = []
                add_info = dict()

                for hostname, edge_hostname in zipped_hostnames:
                    try:
                        host, zone = re.split(r"\.(?=[^.]+\.[^.]+$)", edge_hostname)
                    except:
                        continue

                    df_flt = df_edgehostnames[
                        (df_edgehostnames["recordName"] == host.lower())
                        & (df_edgehostnames["dnsZone"] == zone)
                    ]
                    details.append(df_flt)

                if details:
                    df_flt = pd.concat(details)
                    df_flt.reset_index(drop=True, inplace=True)
                    checked_slots = []

                    for i in df_flt.index:
                        notes = []
                        for col in [
                            "slotNumber",
                            "ipVersionBehavior",
                            "map",
                            "productId",
                        ]:
                            if (col in df_flt.columns) and (
                                str(df_flt.at[i, col]).lower()
                                not in ["none", "nan", ""]
                            ):
                                notes.append(str(df_flt.at[i, col]))

                        add_info[
                            f"{df_flt.at[i,'recordName']}.{df_flt.at[i,'dnsZone']}"
                        ] = ", ".join(notes)

                        # Slot details
                        try:
                            slot = int(df_flt.at[i, "slotNumber"])
                        except:
                            continue

                        if slot in checked_slots:
                            continue
                        else:
                            checked_slots.append(slot)

                        try:
                            df_enroll_flt = df_enroll[
                                df_enroll["assignedSlots"].apply(
                                    lambda x: int(slot) in x
                                )
                            ]

                            for i in df_enroll_flt.index:
                                enroll_id = df_enroll_flt.at[i, "id"]
                                valid_type = df_enroll_flt.at[
                                    i, "validationType"
                                ].upper()
                                cert_type = df_enroll_flt.at[
                                    i, "certificateType"
                                ].upper()
                                valid_type = (
                                    f"{valid_type} {cert_type}"
                                    if valid_type != cert_type
                                    else valid_type
                                )

                                cps_deploy = list_cert_deployments(
                                    switch_key, enroll_id
                                )
                                disallowed_tls = cps_deploy[PROPERTY_VER][
                                    "networkConfiguration"
                                ].get("disallowedTlsVersions")
                                cipher = cps_deploy[PROPERTY_VER][
                                    "networkConfiguration"
                                ].get("mustHaveCiphers")
                                sni = cps_deploy[PROPERTY_VER][
                                    "networkConfiguration"
                                ].get("sniOnly")
                                deployed_nw = cps_deploy[PROPERTY_VER][
                                    "networkConfiguration"
                                ].get("secureNetwork")
                                cert = cps_deploy[PROPERTY_VER][
                                    "primaryCertificate"
                                ].get("certificate")
                                serial = (
                                    parse_certificate(cert)["Serial"] if cert else None
                                )

                                slots.append(
                                    f"{slot}({valid_type}, {'SNI' if sni else 'VIP'}, Disabled TLSv:{('/'.join(i.lower().replace('tlsv','').replace('_','.') for i in disallowed_tls)) if disallowed_tls else None}, {cipher}, {deployed_nw}, Serial:{serial})"
                                )

                                if not all(
                                    [v in disallowed_tls for v in ["TLSv1", "TLSv1_1"]]
                                ):
                                    comments["slots"].append(
                                        "Old TLS versions are allowed"
                                    )
                                if not sni:
                                    comments["slots"].append("VIP Slot is used")
                                if not any(
                                    [
                                        p in cipher
                                        for p in [
                                            "ak-akamai-2020q1",
                                            "ak-akamai-2018q3",
                                        ]
                                    ]
                                ):
                                    comments["slots"].append(
                                        "Non-PFS Cipher profile is used"
                                    )
                        except:
                            slots.append(f"{slot}(No details due to API error)")
                            print(f"Error: {property_name}(Slot:{slot})")

                property_data["edgeHostnames"] = "\n".join(
                    [
                        (
                            f"{host} > {edge_host} ({add_info.get(edge_host)})"
                            if add_info.get(edge_host)
                            else f"{host} > {edge_host}"
                        )
                        for host, edge_host in zipped_hostnames
                    ]
                )
            else:
                property_data["edgeHostnames"] = False

            property_data["slots"] = "\n".join(slots)

            # Server certs
            if server_cert:
                server_cert_results = dict()
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future_to_hostname = {
                        executor.submit(get_cert, hostname): hostname
                        for hostname in df_hostnames["cnameFrom"]
                    }
                    for future in concurrent.futures.as_completed(future_to_hostname):
                        hostname = future_to_hostname[future]
                        try:
                            cert = future.result(timeout=20)
                            server_cert_results[hostname] = cert
                        except:
                            server_cert_results[hostname] = None

                real_certs = []
                for hostname in df_hostnames["cnameFrom"]:
                    try:
                        real_certs.append(
                            f"{hostname} > "
                            + ", ".join(
                                [
                                    f"{k}:{'/'.join(v) if type(v) is list else v}"
                                    for k, v in server_cert_results[hostname].items()
                                ]
                            )
                        )
                        if (
                            property_data["slots"]
                            and (cert["Serial"] not in property_data["slots"])
                            and all(
                                [
                                    (d not in hostname)
                                    for d in [".akamaized.net", ".akamaihd.net"]
                                ]
                            )
                        ):
                            comments["realCertificates"].append(
                                f"Real certificate doesn't match with configuration on {hostname}"
                            )
                    except:
                        real_certs.append(f"{hostname} > None")
                property_data["realCertificates"] = "\n".join(real_certs)

            # Rules
            property_json = get_property_json(
                switch_key,
                property_id,
                version,
            )
            pm_rules = property_json["rules"]
            pm_behaviors = extract_behaviors(pm_rules)

            # Save json file
            directory = os.path.join("property", account)
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            with open(
                os.path.join(
                    directory,
                    f"{property_name}_v{version}.json",
                ),
                mode="w",
                encoding="utf-8_sig",
            ) as f:
                f.write(json.dumps(property_json, indent=4))

            # Origin hostnames
            origin_hostnames = extract_origin_hostnames(pm_rules)
            property_data["originHostnames"] = "\n".join(
                [hostname for hostname in set(origin_hostnames)]
            )

            for origin_hostname in origin_hostnames:
                try:
                    ipaddress.ip_address(origin_hostname)
                except:
                    pass
                else:
                    comments["originHostnames"].append(
                        "Origin hostname is an IP address"
                    )

            # CP codes
            cpcodes = list(set(extract_cpcodes(pm_rules)))
            im_cpcodes = extract_im_cpcodes(pm_rules)
            cpcodes.extend(list(set(im_cpcodes["derivative"])))
            cpcodes = list(set(cpcodes))
            df_cpcodes = list_cpcodes(switch_key, contract_id, group_id)

            # Traffic
            if traffic_report:
                try:
                    df = get_traffic_stats(
                        switch_key, property_id, start_date, end_date
                    )
                    for name, value in df.iloc[0].items():
                        property_data[name] = value
                except:
                    pass

            # Custom maps for SureRoute
            sr_ssmaps = extract_sr_ssmaps(pm_rules)

            # Included/unincluded Behaviors
            available_behaviors = list_available_behaviors(
                switch_key, property_id, version
            )
            product_id = available_behaviors["product"]
            property_data["product"] = product_id

            included_behaviors = sorted(
                set([behavior["name"] for behavior in pm_behaviors])
            )
            not_included_behaviors = [
                behavior
                for behavior in available_behaviors["behaviors"]
                if behavior not in included_behaviors
            ]
            critical_behaviors = [
                behavior
                for behavior in CRITICAL_BEHAVIORS.get(product_id, [])
                if behavior in not_included_behaviors
            ]
            recommended_behaviors = [
                behavior
                for behavior in RECOMMENDED_BEHAVIORS.get(product_id, [])
                if behavior in not_included_behaviors
            ]

            # Conflict behaviors
            conflict_behaviors = []
            for behavior, conflict_behavior in [
                ("tieredDistribution", "siteShield"),
                ("siteShield", "tieredDistribution"),
            ]:
                if behavior in included_behaviors:
                    conflict_behaviors.append(conflict_behavior)
                    critical_behaviors = [
                        b for b in critical_behaviors if b != conflict_behavior
                    ]
                    recommended_behaviors = [
                        b for b in recommended_behaviors if b != conflict_behavior
                    ]

            # Exceptional behaviors
            for behavior in ["persistentConnection"]:
                if behavior not in included_behaviors:
                    critical_behaviors = [
                        b for b in critical_behaviors if b != behavior
                    ]
                    recommended_behaviors = [
                        b for b in recommended_behaviors if b != behavior
                    ]

            property_data["includedBehaviors"] = "\n".join(included_behaviors)
            property_data["notIncludedBehaviors"] = "\n".join(not_included_behaviors)
            property_data["criticalBehaviors"] = "\n".join(critical_behaviors)
            property_data["recomendedBehaviors"] = "\n".join(recommended_behaviors)

            # Error in Property Manager
            if property_json.get("errors"):
                property_data["errors"] = "\n".join(
                    [
                        e["detail"]
                        for e in property_json["errors"]
                        if "detail" in e.keys()
                    ]
                )
            else:
                property_data["errors"] = None

            # Verification note
            property_data["verificationNote"] = None

            # Basic behavior list
            key_behaviors = COLUMN_ORDER.copy()
            key_behaviors.extend(
                [
                    behavior
                    for behavior in sorted(
                        set(
                            CRITICAL_BEHAVIORS.get(product_id, [])
                            + RECOMMENDED_BEHAVIORS.get(product_id, [])
                        )
                    )
                    if behavior not in key_behaviors
                ]
            )
            for key in key_behaviors:
                property_data[key] = None
                comments[key] = []

            # Advanced override
            if pm_rules.get("advancedOverride"):
                xml = pm_rules.get("advancedOverride")
                features = check_xml(xml, available_behaviors["behaviors"])
                if features:
                    property_data["advancedOverride"] = "/".join(features)
                    comments["advancedOverride"].append(
                        f"Advanced override includes {'/'.join(features)}"
                    )
                else:
                    property_data["advancedOverride"] = True

            # Custom override
            if pm_rules.get("customOverride"):
                override_id = pm_rules["customOverride"].get("overrideId")
                property_data["customOverride"] = override_id
                if df_custom_overrides is not None:
                    df_flt = df_custom_overrides[
                        df_custom_overrides["overrideId"] == override_id
                    ]
                    df_flt.reset_index(drop=False, inplace=True)
                    description = df_flt.at[0, "description"]
                    xml = df_flt.at[0, "xml"]
                    features = check_xml(xml, available_behaviors["behaviors"])
                    if features:
                        property_data["customOverride"] = (
                            f"{override_id}({'/'.join(features)})"
                        )
                        comments["customOverride"].append(
                            f"Custom override includes {'/'.join(features)}"
                        )

            # Config check
            for behavior in pm_behaviors:
                behavior_name = behavior["name"]
                behavior_summary = None

                # CP Code
                if behavior_name == "cpCode":
                    behavior_summary = behavior["options"]["value"]["id"]
                    try:
                        cpcode_products = list(
                            df_cpcodes[
                                df_cpcodes["cpcodeId"].astype(int)
                                == int(behavior_summary)
                            ]["productIds"]
                            .reset_index(drop=True)
                            .iloc[0]
                        )
                        behavior_summary = (
                            f"{behavior_summary}({', '.join(cpcode_products)})"
                        )
                        if product_id not in cpcode_products:
                            comments[behavior_name].append(
                                "CP code's product is incorrect"
                            )
                    except:
                        pass

                # SureRoute
                elif behavior_name == "sureRoute":
                    behavior_summary = behavior["options"].get("testObjectUrl", False)

                    # Check Custom map
                    if behavior_summary and sr_ssmaps:
                        sr_ssmap = behavior["options"].get("customMap", False)
                        if sr_ssmap:
                            behavior_summary += f"({sr_ssmap})"
                            sr_ssmap = sr_ssmap in sr_ssmaps
                        if not sr_ssmap:
                            comments[behavior_name].append(
                                "Custom map is not used for SureRoute"
                            )
                    # Check multiple behaviors
                    if (
                        len(
                            [
                                b["name"]
                                for b in pm_behaviors
                                if (
                                    b["name"] == "sureRoute"
                                    and b["options"].get("enabled", False)
                                )
                            ]
                        )
                        > 1
                    ):
                        comments[behavior_name].append(
                            "Multiple SureRoute behaviors are included"
                        )

                # Site Shield
                elif behavior_name == "siteShield":
                    behavior_summary = behavior["options"]["ssmap"].get("value", False)

                # Origin
                elif behavior_name == "origin":
                    behavior_summary = behavior["options"].get("originType")

                    if behavior_summary == "NET_STORAGE":
                        behavior_summary = behavior["options"]["netStorage"].get(
                            "downloadDomainName"
                        )
                    elif behavior_summary == "MEDIA_SERVICE_LIVE":
                        behavior_summary = behavior["options"].get("mslorigin")
                    elif behavior_summary == "CUSTOMER":
                        behavior_summary = behavior["options"].get("hostname")
                        details = []

                        # Foward Host Header
                        if behavior["options"].get("customForwardHostHeader"):
                            details.append(
                                f"FHH:{behavior['options']['customForwardHostHeader']}"
                            )
                        elif behavior["options"].get("forwardHostHeader"):
                            details.append(
                                f"FHH:{behavior['options']['forwardHostHeader']}"
                            )

                        # Cache Key Hostname
                        if behavior["options"].get("cacheKeyHostname"):
                            details.append(
                                f"CK:{behavior['options']['cacheKeyHostname']}"
                            )

                        # SNI
                        if behavior["options"].get("originSni") is not None:
                            details.append(f"SNI:{behavior['options']['originSni']}")

                        # IP version
                        if behavior["options"].get("ipVersion") is not None:
                            details.append(behavior["options"]["ipVersion"])

                        # Trust cert
                        if (
                            behavior["options"].get("originCertsToHonor")
                            == "CUSTOM_CERTIFICATES"
                        ):
                            details.append("Pinning")

                        behavior_summary += f"({', '.join([v for v in details if v])})"

                    if behavior["options"].get("cacheKeyHostname") not in [
                        "REQUEST_HOST_HEADER"
                    ]:
                        comments[behavior_name].append(
                            "Cache key hostname is not incoming host header"
                        )
                    if behavior["options"].get("originSni") is False:
                        comments[behavior_name].append("Origin SNI is disabled")
                    if (
                        behavior["options"].get("originCertsToHonor")
                        == "CUSTOM_CERTIFICATES"
                    ):
                        comments[behavior_name].append(
                            "Origin server certificate is pinned"
                        )

                # Persistent connection
                elif behavior_name == "persistentConnection":
                    behavior_summary = str(behavior["options"].get("timeout", False))

                # SiteFailover
                elif behavior_name == "failAction":
                    behavior_summary = behavior["options"].get("actionType", False)
                    if behavior_summary:
                        for key in behavior["options"].keys():
                            if key in ["netStorageHostname"]:
                                behavior_summary += f"({behavior['options'][key].get('downloadDomainName')})"
                                break
                            elif re.search(r"hostname$", key.lower()):
                                behavior_summary += f"({behavior['options'][key]})"
                                break

                # Enhanced Akamai Protocol
                elif behavior_name == "enhancedAkamaiProtocol":
                    behavior_summary = True

                # HTTP2
                elif behavior_name == "http2":
                    behavior_summary = True
                    # Check Chunked Transfer Encoding
                    if not any(
                        [
                            b["options"].get("enabled", False)
                            for b in pm_behaviors
                            if b["name"] == "allowTransferEncoding"
                        ]
                    ):
                        comments["allowTransferEncoding"].append(
                            "'Chunked Transfer Encoding' is not enabled"
                        )

                # Last Mile Acceleration (gzip)
                elif behavior_name == "gzipResponse":
                    behavior_summary = behavior["options"].get("behavior")
                    if behavior_summary:
                        if behavior_summary != "ALWAYS":
                            comments[behavior_name].append(
                                "'Last Mile Acceleration' is not ALWAYS"
                            )
                        if "contentType is" not in behavior["criteria"]:
                            comments[behavior_name].append(
                                "'Last Mile Acceleration' criteria is not Content-Type"
                            )

                # Caching
                elif behavior_name == "caching":
                    if behavior["options"].get("ttl"):
                        behavior_summary = behavior["options"]["ttl"]
                    else:
                        behavior_summary = behavior["options"].get("behavior", False)

                # Image and Video Manager
                elif behavior_name in ["imageManager", "imageManagerVideo"]:
                    behavior_summary = behavior["options"].get("enabled")
                    if behavior_summary:
                        for key in behavior["options"].keys():
                            if re.match(r"policy", key.lower()):
                                im_cpcode = "/".join(
                                    [
                                        f"{behavior['options'][tag]['id']}"
                                        for tag in [
                                            "cpCodeOriginal",
                                            "cpCodeTransformed",
                                        ]
                                    ]
                                )
                                behavior_summary = (
                                    f"{behavior['options'][key]}(CP code:{im_cpcode})"
                                )
                                break

                # Origin/Client Characteristics
                elif behavior_name in [
                    "originCharacteristics",
                    "clientCharacteristics",
                ]:
                    behavior_summary = behavior["options"].get("country", False)

                # Set variable
                elif behavior["name"] == "setVariable":
                    behavior_summary = behavior["options"].get("variableName", False)

                # Log delivery
                elif behavior["name"] == "report":
                    behavior_summary = "/".join(
                        [
                            k.replace("log", "")
                            for k, v in behavior["options"].items()
                            if "log" in str(k)
                            and str(v).lower() not in ["false", "off"]
                        ]
                    )

                # Advanced behavior
                elif behavior["name"] == "advanced":
                    xml = behavior["options"].get("xml")
                    features = check_xml(xml, available_behaviors["behaviors"])
                    if features:
                        behavior_summary = "/".join(features)
                        comments[behavior_name].append(
                            f"Advanced behavior includes {'/'.join(features)}"
                        )
                    else:
                        behavior_summary = True

                # Custom behavior
                elif behavior["name"] == "customBehavior":
                    behavior_id = behavior["options"].get("behaviorId")
                    behavior_summary = behavior_id
                    if df_custom_behaviors is not None:
                        df_flt = df_custom_behaviors[
                            df_custom_behaviors["behaviorId"] == behavior_id
                        ]
                        df_flt.reset_index(drop=False, inplace=True)
                        description = df_flt.at[0, "description"]
                        xml = df_flt.at[0, "xml"]
                        features = check_xml(xml, available_behaviors["behaviors"])
                        if features:
                            behavior_summary = f"{behavior_id}({'/'.join(features)})"
                            comments[behavior_name].append(
                                f"Custom behavior includes {'/'.join(features)}"
                            )

                # Others
                elif any(
                    [
                        "enable" in option.lower()
                        for option in behavior["options"].keys()
                    ]
                ):
                    behavior_summary = "/".join(
                        [
                            (
                                re.sub(r"(^enable|Enable$)", "", k)
                                if k not in ["enable", "enabled"]
                                else str(v)
                            )
                            for k, v in behavior["options"].items()
                            if (
                                (
                                    ("enable" in k.lower())
                                    and (str(v).lower() not in ["false", "off"])
                                )
                                or k in ["enable", "enabled"]
                            )
                        ]
                    )

                else:
                    for option in [
                        "behavior",
                        "value",
                        "action",
                        "description",
                    ]:
                        if option in behavior["options"].keys():
                            behavior_summary = behavior["options"].get(option)
                            break
                    if behavior_summary in ["", None]:
                        behavior_summary = ", ".join(
                            [f"{k}:{v}" for k, v in behavior["options"].items()]
                        )

                # Highlight disabled behaviors
                if behavior_summary is False:
                    comments[behavior_name].append(f"{behavior_name} is disabled")

                # Check Allow All Methods on Parent Servers when SR/TD/SS is enabled even if not configured
                if behavior_name in [
                    "sureRoute",
                    "tieredDistribution",
                    "siteShield",
                ] and (behavior_summary not in [False, None]):
                    if not any(
                        [
                            b["options"].get("enabled", False)
                            for b in pm_behaviors
                            if b["name"] == "allHttpInCacheHierarchy"
                        ]
                    ):
                        comments["allHttpInCacheHierarchy"].append(
                            "'Allow All Methods on Parent Servers' is not enabled"
                        )

                # Write result to dict
                if behavior_summary is not None:
                    if behavior_name not in property_data.keys():
                        property_data[behavior_name] = None

                    behavior_summary = f"{behavior_summary} @{behavior['rule_name']}"

                    if property_data[behavior_name] is None:
                        property_data[behavior_name] = str(behavior_summary)
                    elif behavior_summary not in property_data[behavior_name]:
                        property_data[behavior_name] += f"\n{behavior_summary}"

            # Ceck Advanced/Custom configs
            for behavior in available_behaviors["behaviors"]:
                for col in [
                    "advanced",
                    "customBehavior",
                    "advancedOverride",
                    "customOverride",
                ]:
                    if property_data[col]:
                        for line in str(property_data[col]).split("\n"):
                            if behavior in line:
                                text = (
                                    f"Configured by {col}:{line}"
                                    if col in ["advanced", "customBehavior"]
                                    else f"Configured by {col}"
                                )
                                property_data[behavior] = (
                                    text
                                    if not property_data[behavior]
                                    else f"{property_data[behavior]}\n{text}"
                                )

            # Verification note
            property_data["verificationNote"] = "\n".join(
                sorted(set(sum(comments.values(), [])))
            )

            # Disabled critical/recommended behaviors
            for behavior in sorted(
                set(
                    CRITICAL_BEHAVIORS.get(product_id, [])
                    + RECOMMENDED_BEHAVIORS.get(product_id, [])
                )
            ):
                if (property_data[behavior] is not None) and (
                    behavior not in conflict_behaviors
                ):
                    if all(
                        [
                            re.sub(r"\@.*$", "", value).strip().lower()
                            in ["false", "never"]
                            for value in property_data[behavior].split("\n")
                        ]
                    ):
                        if behavior in CRITICAL_BEHAVIORS.get(product_id, []):
                            critical_behaviors.append(behavior)
                            critical_behaviors.sort()
                        if behavior in RECOMMENDED_BEHAVIORS.get(product_id, []):
                            recommended_behaviors.append(behavior)
                            recommended_behaviors.sort()

            property_data["criticalBehaviors"] = "\n".join(
                sorted(set(critical_behaviors))
            )
            property_data["recomendedBehaviors"] = "\n".join(
                sorted(set(recommended_behaviors))
            )

            return pd.DataFrame([property_data]), {property_name: comments}
        except Exception as e:
            if retry_counter == 0:
                print(f"Error: {e} @{property_name}")
                property_data = {
                    "accountName": account,
                    "accountSwitchKey": switch_key,
                    "propertyName": property_name,
                }
                return pd.DataFrame([property_data]), dict()
            else:
                retry_counter -= 1
                time.sleep(30)




def save_df_to_excel(df, file_xlsx, mode, sheet_name):
    def clean_illegal_characters(value):
        if isinstance(value, str):
            return "".join(
                char for char in value if ord(char) >= 32 or char in "\t\n\r"
            )
        return value

    df = df.map(clean_illegal_characters)
    with pd.ExcelWriter(
        file_xlsx,
        mode=mode,
        engine="openpyxl",
        if_sheet_exists="replace" if mode == "a" else None,
    ) as f:
        df.to_excel(
            f,
            sheet_name=sheet_name,
            index=False,
        )

        try:
            # Add border lines
            workbook = f.book
            worksheet = workbook[sheet_name]

            for i in df.index.values:
                set_border = False
                if i == len(df) - 1:
                    set_border = True
                elif df.iloc[i, 0] != df.iloc[i + 1, 0]:
                    set_border = True

                if set_border:
                    for j in range(df.shape[1]):
                        cell = worksheet.cell(row=i + 2, column=j + 1)
                        cell.border = Border(bottom=Side(style="thin"))

            # Highlight cells
            for i in df.index.values:
                comments = all_comments[df.at[i, "propertyName"]]

                for j, col in enumerate(df.columns.values):
                    cell = worksheet.cell(row=i + 2, column=j + 1)
                    fill_color = False

                    if (str(df.iloc[i, j]).lower() != "true") and (
                        col in comments.keys()
                    ):
                        fill_color = bool(comments[col])
                    elif col in [
                        "criticalBehaviors",
                        "recomendedBehaviors",
                        "verificationNote",
                        "errors",
                        "warnings",
                    ]:
                        fill_color = bool(df.iloc[i, j])

                    if fill_color:
                        cell.fill = PatternFill(
                            start_color="FFC0CB",
                            end_color="FFC0CB",
                            fill_type="solid",
                        )
        except:
            pass


if __name__ == "__main__":
    try:
        # Clean and create/recreate folders
        for folder in ["output", "merge_details"]:
            if os.path.exists(folder):
                shutil.rmtree(folder)
            os.makedirs(folder)

        if ACCOUNTS:
            for i, account in enumerate(ACCOUNTS):
                print(f"[{i+1:3d}] {account}")

        file_xlsx = None
        mode = "w"
        n_account = 0

        today = date.today()
        months = 1
        start_date = datetime(
            today.year if (today.month - months) > 0 else (today.year - 1),
            (today.month - months) % 12 if (today.month - months) % 12 else 12,
            1,
            tzinfo=dateutil.tz.gettz("UTC"),
        )
        end_date = datetime(today.year, today.month, 1, tzinfo=dateutil.tz.gettz("UTC"))

        check_all_property = (
            input(f"\nCheck all properties in {PROPERTY_VER}? (y/n):").lower() == "y"
        )
        traffic_report = (
            input("Traffic data(Last month) are required? (y/n):").lower() == "y"
        )
        server_cert = input("Real certificates are required? (y/n):").lower() == "y"

        while True:
            n_account += 1
            print(f"\n------------ Account #{n_account} ------------")

            if ACCOUNTS:
                account = ACCOUNTS.pop(0)
            else:
                account = input("\nAccount name:").strip()
                if not account:
                    break

            if not file_xlsx:
                file_xlsx = (
                    f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_properties.xlsx"
                )

            # SW key
            df = get_switch_keys(account)
            data_dict = extract_row_of_dataframe(df, display_table=True)
            switch_key = data_dict.get("accountSwitchKey")
            account = re.sub(r'([\/:*?"<>|[\]]|_.+$)', "", data_dict.get("accountName"))
            sheet_name = account[:31]

            # Properties
            try:
                df_group = list_groups(switch_key, parent_only=False)
                groups = []
                properties = []

                print("\n[Groups]")
                try:
                    display(df_group)
                except:
                    print(df_group)

                for i in tqdm(df_group.index):
                    for contract_id in df_group.at[i, "contractIds"]:
                        group_id = df_group.at[i, "groupId"]
                        df = list_properties(
                            switch_key,
                            contract_id,
                            group_id,
                            network=PROPERTY_VER,
                        )
                        if (df is not None) and len(df):
                            properties.append(df)

                properties = pd.concat(properties)
                properties = sorted(set(properties["propertyName"]))
            except:
                print(f"No Property in {PROPERTY_VER}")
                continue

            print(
                f"\n[Properties in {PROPERTY_VER} ({len(properties)})]\n{', '.join(properties)}"
            )

            if (not check_all_property) and (len(properties) > 1):
                temp_list = input(
                    f"Property name(s) (If multiple properties, separate them by comma, e.g., {', '.join(properties[:2])}):"
                ).strip()
                if not temp_list:
                    properties = []
                else:
                    temp_list = [
                        p.strip().lower() for p in re.split(r"[,\s]+", temp_list)
                    ]
                    temp_list = [p for p in properties if p.lower() in temp_list]
                    properties = temp_list

            # Edge hostnames
            df_edgehostnames = list_edgehostnames(
                switch_key,
            )

            # Slots
            df_enroll = get_cps_enrollment_id(switch_key)

            # Custom behaviors/overrides
            df_custom_behaviors = list_custom_behaviors(switch_key)
            df_custom_overrides = list_custom_overrides(switch_key)

            result = []
            all_comments = dict()

            print(f"\nProcessing (Total: {int(len(properties)/N_CHUNK)+1}it)")
            with concurrent.futures.ThreadPoolExecutor() as executor:
                for chunk in tqdm(chunked_list(properties, N_CHUNK)):
                    futures = [
                        executor.submit(
                            property_analysis,
                            switch_key,
                            property_name,
                            df_enroll,
                            df_custom_behaviors=df_custom_behaviors,
                            df_custom_overrides=df_custom_overrides,
                            target_ver=PROPERTY_VER,
                            traffic_report=traffic_report,
                            server_cert=server_cert,
                        )
                        for property_name in chunk
                    ]
                    for future in concurrent.futures.as_completed(futures):
                        df, comments = future.result()
                        result.append(df)
                        all_comments.update(comments)

                    # Save as Excel file
                    #if len(result):
                    #    df = pd.concat(result, ignore_index=True)
                    #    save_df_to_excel(df, file_xlsx, mode, sheet_name)
                    #    mode = "a"
                    
                    if len(result):
                        df = pd.concat(result, ignore_index=True)
                        csv_filename = f"output/{account}.csv".replace(" ", "_")
                        df.to_csv(csv_filename, index=False)
                        #print(f"✅ Saved CSV for {account}: {csv_filename}")
            print(f"✅ Saved CSV for {account}: {csv_filename}")

        # create a merged report
        # Ask user if they want to create a combined report
        combine = input("\nDo you want to create a combined report? (y/n): ").strip().lower()
        if combine == "y":
            merge_csv = input("Enter path to CSV file listing account names to merge: ").strip()

            selected_accounts = []
            try:
                with open(merge_csv, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    for row in reader:
                        if row and row[0].strip().lower() != "account":
                            selected_accounts.append(row[0].strip().replace(" ", "_"))
            except Exception as e:
                print(f"❌ Error reading merge list: {e}")
                selected_accounts = []

            merged_df = pd.DataFrame()
            for account in selected_accounts:
                file_path = os.path.join("output", f"{account}.csv")
                if os.path.exists(file_path):
                    df = pd.read_csv(file_path)
                    df["source_account"] = account  # optional: tag source
                    merged_df = pd.concat([merged_df, df], ignore_index=True)
                else:
                    print(f"⚠️  CSV not found for account: {account}")

            if not merged_df.empty:
                output_file = os.path.join("merge_details", "merge_details.csv")
                merged_df.to_csv(output_file, index=False)
                print(f"\n✅ Combined report saved to: {output_file}")
            else:
                print("❌ No matching files found. No combined report created.")


    except Exception as e:
        print(f"\nError: {e}")